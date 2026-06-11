"""
P1dB Compression Point Test Automation Software
================================================
Professional PyQt6 desktop application for automating P1dB Compression Point testing
using Keysight Signal Generator and Signal Analyzer instruments.

Architecture:
    - InstrumentManager  : VISA detection, connection, SCPI abstraction
    - TestEngine         : Sweep logic, P1dB detection, runs in QThread
    - ExcelReporter      : OpenPyXL report generation
    - MainWindow         : PyQt6 GUI, thread-safe signal/slot wiring
    - ThemeManager       : Dark / Light theme stylesheets

Author  : Generated for production RF test environments
Requires: Python 3.10+, PyQt6, pyvisa, openpyxl
"""

# ---------------------------------------------------------------------------
# Standard Library
# ---------------------------------------------------------------------------
import sys
import os
import time
import logging
import traceback
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Third-Party
# ---------------------------------------------------------------------------
try:
    import pyvisa
except ImportError:
    pyvisa = None  # Graceful degradation for UI-only preview

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QGroupBox, QLabel, QLineEdit, QPushButton, QComboBox,
    QCheckBox, QProgressBar, QTextEdit, QStatusBar, QFrame, QSplitter,
    QScrollArea, QMessageBox, QFileDialog, QTabWidget, QSizePolicy,
    QSpacerItem, QDoubleSpinBox, QSpinBox
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, pyqtSlot, QTimer, QMutex, QMutexLocker,
    QSettings
)
from PyQt6.QtGui import QFont, QColor, QPalette, QIcon, QTextCursor

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------
LOG_DIR = Path("C:/P1dB_Test_Results/logs")
LOG_DIR.mkdir(parents=True, exist_ok=True)
_log_file = LOG_DIR / f"p1db_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(_log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("P1dB")


# ===========================================================================
# INSTRUMENT MANAGER
# ===========================================================================

class InstrumentError(Exception):
    """Raised for instrument communication failures."""


class Instrument:
    """
    Thin SCPI wrapper around a single pyvisa Resource.
    Stores identity fields and provides query / write helpers.
    """

    def __init__(self, resource, address: str):
        self.resource = resource
        self.address = address
        self.manufacturer = ""
        self.model = ""
        self.serial = ""
        self.firmware = ""
        self.conn_type = ""  # USB / LAN / GPIB

    # ------------------------------------------------------------------
    def query(self, cmd: str, timeout_ms: int = 10_000) -> str:
        self.resource.timeout = timeout_ms
        try:
            resp = self.resource.query(cmd).strip()
            logger.debug("QUERY  [%s] %r -> %r", self.address, cmd, resp)
            return resp
        except Exception as exc:
            raise InstrumentError(f"Query failed ({cmd}): {exc}") from exc

    def write(self, cmd: str) -> None:
        try:
            self.resource.write(cmd)
            logger.debug("WRITE  [%s] %r", self.address, cmd)
        except Exception as exc:
            raise InstrumentError(f"Write failed ({cmd}): {exc}") from exc

    def identify(self) -> None:
        """Populate manufacturer / model / serial / firmware from *IDN?."""
        idn = self.query("*IDN?")
        parts = [p.strip() for p in idn.split(",")]
        self.manufacturer = parts[0] if len(parts) > 0 else "Unknown"
        self.model        = parts[1] if len(parts) > 1 else "Unknown"
        self.serial       = parts[2] if len(parts) > 2 else "Unknown"
        self.firmware     = parts[3] if len(parts) > 3 else "Unknown"
        # Derive connection type from address string
        addr_upper = self.address.upper()
        if "USB" in addr_upper:
            self.conn_type = "USB"
        elif "TCPIP" in addr_upper or "SOCKET" in addr_upper:
            self.conn_type = "LAN"
        elif "GPIB" in addr_upper:
            self.conn_type = "GPIB"
        else:
            self.conn_type = "Other"

    def close(self) -> None:
        try:
            self.resource.close()
        except Exception:
            pass

    def __repr__(self):
        return f"<Instrument {self.model} @ {self.address}>"


# ---------------------------------------------------------------------------

class InstrumentManager:
    """
    Discovers and classifies Keysight instruments over USB / LAN / Both.
    Instruments are classified as Signal Generator or Signal Analyzer based
    on their *IDN? model string — no hardcoded model numbers.
    """

    # Keywords that appear in Keysight signal generator model strings
    _SG_KEYWORDS = ("E44", "MXG", "EXG", "CXG", "N5171", "N5172", "N5173",
                    "N5181", "N5182", "N5183", "E8257", "E8267", "PSG", "ESG",
                    "N9310", "33600", "33500", "33522", "33512")

    # Keywords that appear in Keysight signal / spectrum analyzer model strings
    _SA_KEYWORDS = ("N9020", "N9030", "N9040", "N9041", "N9068", "MXA", "PXA",
                    "EXA", "CXA", "N9000", "E4440", "E4443", "E4445", "E4446",
                    "E4447", "E4448", "PSA", "ESA", "N9010", "N9038",
                    "FieldFox", "E4406", "VSA")

    def __init__(self, filter_mode: str = "AUTO"):
        """
        filter_mode: 'AUTO' | 'USB' | 'LAN'
        """
        self.filter_mode = filter_mode.upper()
        self._rm: pyvisa.ResourceManager | None = None
        self.signal_generator: Instrument | None = None
        self.signal_analyzer:  Instrument | None = None

    # ------------------------------------------------------------------
    def _open_rm(self) -> pyvisa.ResourceManager:
        if pyvisa is None:
            raise InstrumentError("PyVISA is not installed.")
        if self._rm is None:
            self._rm = pyvisa.ResourceManager()
        return self._rm

    def _address_passes_filter(self, addr: str) -> bool:
        upper = addr.upper()
        if self.filter_mode == "USB":
            return "USB" in upper
        if self.filter_mode == "LAN":
            return "TCPIP" in upper or "SOCKET" in upper
        return True  # AUTO → accept all

    @staticmethod
    def _is_signal_generator(model: str) -> bool:
        m = model.upper()
        return any(k.upper() in m for k in InstrumentManager._SG_KEYWORDS)

    @staticmethod
    def _is_signal_analyzer(model: str) -> bool:
        m = model.upper()
        return any(k.upper() in m for k in InstrumentManager._SA_KEYWORDS)

    # ------------------------------------------------------------------
    def discover(self, progress_cb=None) -> tuple[list[str], list[str]]:
        """
        Scan VISA resources and return (sg_addresses, sa_addresses).
        progress_cb(message: str) is called with status updates.
        """
        def _cb(msg):
            if progress_cb:
                progress_cb(msg)
            logger.info(msg)

        rm = self._open_rm()
        resources = rm.list_resources()
        _cb(f"Found {len(resources)} VISA resource(s). Filtering…")

        sg_list, sa_list = [], []

        for addr in resources:
            if not self._address_passes_filter(addr):
                continue
            _cb(f"Probing {addr} …")
            try:
                res = rm.open_resource(addr)
                res.timeout = 3000
                idn = res.query("*IDN?").strip()
                parts = [p.strip() for p in idn.split(",")]
                manufacturer = parts[0] if parts else ""
                model = parts[1] if len(parts) > 1 else ""

                if "KEYSIGHT" not in manufacturer.upper() and \
                   "AGILENT"  not in manufacturer.upper() and \
                   "HEWLETT"  not in manufacturer.upper():
                    res.close()
                    continue

                if self._is_signal_generator(model):
                    sg_list.append(addr)
                    _cb(f"  → Signal Generator: {model}")
                elif self._is_signal_analyzer(model):
                    sa_list.append(addr)
                    _cb(f"  → Signal Analyzer:  {model}")
                else:
                    _cb(f"  → Keysight instrument (unclassified): {model}")
                res.close()
            except Exception as exc:
                _cb(f"  → Could not identify {addr}: {exc}")

        return sg_list, sa_list

    def connect_sg(self, address: str) -> Instrument:
        rm = self._open_rm()
        res = rm.open_resource(address)
        inst = Instrument(res, address)
        inst.identify()
        self.signal_generator = inst
        return inst

    def connect_sa(self, address: str) -> Instrument:
        rm = self._open_rm()
        res = rm.open_resource(address)
        inst = Instrument(res, address)
        inst.identify()
        self.signal_analyzer = inst
        return inst

    def disconnect_all(self) -> None:
        for inst in (self.signal_generator, self.signal_analyzer):
            if inst:
                inst.close()
        self.signal_generator = None
        self.signal_analyzer  = None
        if self._rm:
            try:
                self._rm.close()
            except Exception:
                pass
            self._rm = None


# ===========================================================================
# SCPI HELPERS  — Signal Generator & Signal Analyzer
# ===========================================================================

class SignalGeneratorSCPI:
    """High-level SCPI commands for Keysight signal generators."""

    def __init__(self, inst: Instrument):
        self.inst = inst

    def reset(self):
        self.inst.write("*RST")
        time.sleep(0.5)

    def set_frequency_hz(self, freq_hz: float):
        self.inst.write(f":FREQ {freq_hz:.6f}Hz")

    def set_power_dbm(self, power_dbm: float):
        self.inst.write(f":POW {power_dbm:.3f}dBm")

    def rf_on(self):
        self.inst.write(":OUTP ON")

    def rf_off(self):
        self.inst.write(":OUTP OFF")

    def get_power_dbm(self) -> float:
        return float(self.inst.query(":POW?"))

    def get_frequency_hz(self) -> float:
        return float(self.inst.query(":FREQ?"))


class SignalAnalyzerSCPI:
    """High-level SCPI commands for Keysight signal / spectrum analyzers."""

    def __init__(self, inst: Instrument):
        self.inst = inst

    def reset(self):
        self.inst.write("*RST")
        time.sleep(1.0)

    def set_frequency_hz(self, freq_hz: float):
        self.inst.write(f":SENS:FREQ:CENT {freq_hz:.6f}Hz")

    def set_reference_level(self, ref_dbm: float):
        self.inst.write(f":DISP:WIND:TRAC:Y:RLEV {ref_dbm:.2f}dBm")

    # ---- Auto Tune -------------------------------------------------------
    def auto_tune(self) -> dict:
        """
        Execute Auto Tune sequence. Returns dict of current analyzer settings.
        Equivalent of pressing Auto Tune on the front panel.
        """
        # Signal Analyzers expose auto-tune as :SENS:FREQuency:TUNE:IMMediate
        # or via the coupling commands — we cover the most common variants.
        try:
            self.inst.write(":SENS:FREQ:TUNE:IMM")
        except Exception:
            pass
        # Also trigger auto-scaling
        try:
            self.inst.write(":DISP:WIND:TRAC:Y:AUTO")
        except Exception:
            pass
        time.sleep(1.0)  # Allow analyzer to settle
        return self._read_settings()

    def _read_settings(self) -> dict:
        def safe_query(cmd, default=0.0):
            try:
                return float(self.inst.query(cmd))
            except Exception:
                return default

        return {
            "center_freq_hz": safe_query(":SENS:FREQ:CENT?"),
            "span_hz":        safe_query(":SENS:FREQ:SPAN?"),
            "rbw_hz":         safe_query(":SENS:BAND:RES?"),
            "vbw_hz":         safe_query(":SENS:BAND:VID?"),
            "ref_level_dbm":  safe_query(":DISP:WIND:TRAC:Y:RLEV?"),
        }

    # ---- Peak Search -----------------------------------------------------
    def peak_search(self) -> tuple[float, float]:
        """
        Execute Peak Search and return (marker_freq_hz, marker_amp_dbm).
        """
        # Activate marker 1 and run peak search
        self.inst.write(":CALC:MARK1:STAT ON")
        self.inst.write(":CALC:MARK1:MAX")
        time.sleep(0.3)
        freq = float(self.inst.query(":CALC:MARK1:X?"))
        amp  = float(self.inst.query(":CALC:MARK1:Y?"))
        return freq, amp

    # ---- Channel Power ---------------------------------------------------
    def configure_channel_power(self, center_hz: float, bw_hz: float):
        self.inst.write(":SENS:POW:ACH:BWID:CHAN1 {:.6f}Hz".format(bw_hz))
        self.inst.write(":CONF:CHP")
        self.set_frequency_hz(center_hz)

    def read_channel_power(self) -> float:
        self.inst.write(":INIT:IMM;*WAI")
        result = self.inst.query(":FETC:CHP:CHP?")
        return float(result)

    # ---- Single sweep ----------------------------------------------------
    def single_sweep(self):
        self.inst.write(":INIT:CONT OFF")
        self.inst.write(":INIT:IMM;*WAI")
        time.sleep(0.3)

    def get_settings(self) -> dict:
        return self._read_settings()


# ===========================================================================
# EXCEL REPORTER
# ===========================================================================

class ExcelReporter:
    """Creates and manages the OpenPyXL workbook for P1dB results."""

    OUTPUT_ROOT = Path("C:/P1dB_Test_Results")

    # Color fills
    _GREEN_FILL  = PatternFill("solid", fgColor="00B050") if openpyxl else None
    _HEADER_FILL = PatternFill("solid", fgColor="1F3864") if openpyxl else None
    _ALT_FILL    = PatternFill("solid", fgColor="D9E1F2") if openpyxl else None
    _WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF") if openpyxl else None

    def __init__(self):
        self.wb = None
        self.ws = None
        self.filepath: Path | None = None
        self._row = 1
        self._data_start_row = 1

    # ------------------------------------------------------------------
    def create(self, meta: dict) -> Path:
        if openpyxl is None:
            raise RuntimeError("openpyxl is not installed.")

        # Build folder path
        proj   = self._sanitize(meta.get("project_name", "Project"))
        unit   = self._sanitize(meta.get("unit_name", "Unit"))
        cond   = self._sanitize(meta.get("test_condition", "Cond"))
        folder = self.OUTPUT_ROOT / f"{proj}_{unit}_{cond}"
        folder.mkdir(parents=True, exist_ok=True)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filepath = folder / f"P1dB_Test_{ts}.xlsx"

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "P1dB Results"

        self._write_header(meta)
        self._write_column_headers()
        self.wb.save(self.filepath)
        logger.info("Excel report created: %s", self.filepath)
        return self.filepath

    @staticmethod
    def _sanitize(s: str) -> str:
        """Remove characters unsafe for folder names."""
        return "".join(c for c in s if c.isalnum() or c in ("_", "-"))[:40]

    def _hdr_font(self):
        return Font(bold=True, color="FFFFFF", size=11)

    def _write_header(self, meta: dict):
        ws = self.ws
        fields = [
            ("Project Name",        meta.get("project_name", "")),
            ("Unit Name",           meta.get("unit_name", "")),
            ("Test Condition",      meta.get("test_condition", "")),
            ("Date",                datetime.now().strftime("%Y-%m-%d")),
            ("Time",                datetime.now().strftime("%H:%M:%S")),
            ("SG Manufacturer",     meta.get("sg_manufacturer", "")),
            ("SG Model",            meta.get("sg_model", "")),
            ("SG Serial Number",    meta.get("sg_serial", "")),
            ("SA Manufacturer",     meta.get("sa_manufacturer", "")),
            ("SA Model",            meta.get("sa_model", "")),
            ("SA Serial Number",    meta.get("sa_serial", "")),
            ("Center Frequency",    meta.get("center_freq_str", "")),
            ("Frequency Unit",      meta.get("freq_unit", "MHz")),
            ("Reference Level",     f"{meta.get('ref_level', 0)} dBm"),
            ("Input Cable Loss",    f"{meta.get('in_loss', 0)} dB"),
            ("Output Cable Loss",   f"{meta.get('out_loss', 0)} dB"),
            ("Measurement Type",    meta.get("meas_type", "Marker Power")),
            ("Span",                meta.get("span_str", "")),
            ("RBW",                 meta.get("rbw_str", "")),
            ("VBW",                 meta.get("vbw_str", "")),
            ("Marker Frequency",    meta.get("marker_freq_str", "")),
            ("Communication Type",  meta.get("conn_type", "AUTO")),
        ]

        row = 1
        for label, value in fields:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, color="1F3864", size=10)
            ws.cell(row=row, column=2, value=str(value)).font = Font(size=10)
            row += 1

        # Spacer row
        self._row = row + 1

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        for col in range(3, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _write_column_headers(self):
        ws = self.ws
        headers = [
            "Timestamp", "Input Power (dBm)", "Output Power (dBm)",
            "Gain (dB)", "Compression (dB)"
        ]
        col_start = 4
        for i, h in enumerate(headers):
            cell = ws.cell(row=self._row, column=col_start + i, value=h)
            cell.fill = self._HEADER_FILL
            cell.font = self._hdr_font()
            cell.alignment = Alignment(horizontal="center")

        self._data_start_row = self._row + 1
        self._row = self._data_start_row

    # ------------------------------------------------------------------
    def append_row(self, ts: str, in_pwr: float, out_pwr: float,
                   gain: float, compression: float, highlight: bool = False):
        if self.ws is None:
            return
        ws = self.ws
        row = self._row
        values = [ts, round(in_pwr, 3), round(out_pwr, 3),
                  round(gain, 3), round(compression, 3)]
        fill = self._GREEN_FILL if highlight else (
            self._ALT_FILL if (row - self._data_start_row) % 2 == 0 else self._WHITE_FILL
        )
        col_start = 4
        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=col_start + i, value=v)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        self._row += 1
        self.wb.save(self.filepath)

    def save(self):
        if self.wb and self.filepath:
            self.wb.save(self.filepath)
            logger.info("Excel saved: %s", self.filepath)


# ===========================================================================
# TEST ENGINE  (QThread)
# ===========================================================================

class TestEngine(QThread):
    """
    Runs the P1dB sweep in a background thread.
    Communicates with the GUI via Qt signals.
    """

    # Signals ---------------------------------------------------------------
    sig_log         = pyqtSignal(str)                    # Activity log message
    sig_status      = pyqtSignal(str)                    # Status bar message
    sig_progress    = pyqtSignal(int, int)               # current, total
    sig_measurement = pyqtSignal(dict)                   # live reading dict
    sig_p1db_found  = pyqtSignal(dict)                   # P1dB result dict
    sig_error       = pyqtSignal(str)                    # Error message
    sig_finished    = pyqtSignal()                       # Sweep finished normally
    sig_tune_data   = pyqtSignal(dict)                   # Auto-tune results
    sig_excel_saved = pyqtSignal(str)                    # Excel file path

    def __init__(self, sg: SignalGeneratorSCPI, sa: SignalAnalyzerSCPI,
                 params: dict, reporter: ExcelReporter):
        super().__init__()
        self.sg = sg
        self.sa = sa
        self.params = params
        self.reporter = reporter
        self._stop_flag = False
        self._mutex = QMutex()

    def stop(self):
        with QMutexLocker(self._mutex):
            self._stop_flag = True

    def _stopped(self) -> bool:
        with QMutexLocker(self._mutex):
            return self._stop_flag

    # ------------------------------------------------------------------
    def run(self):
        p = self.params
        sg = self.sg
        sa = self.sa

        try:
            # ---- Configure instruments --------------------------------
            self.sig_status.emit("Configuring instruments…")
            self.sig_log.emit("Configuring Signal Generator…")
            sg.set_frequency_hz(p["center_freq_hz"])

            self.sig_log.emit("Configuring Signal Analyzer…")
            sa.set_frequency_hz(p["center_freq_hz"])
            sa.set_reference_level(p["ref_level"])

            # ---- RF ON -----------------------------------------------
            self.sig_log.emit("RF ON")
            sg.set_power_dbm(p["start_power"])
            sg.rf_on()
            time.sleep(0.5)

            # ---- Auto Tune -------------------------------------------
            if p.get("auto_tune", True):
                self.sig_log.emit("Executing Auto Tune…")
                self.sig_status.emit("Auto Tune…")
                tune_data = sa.auto_tune()
                self.sig_tune_data.emit(tune_data)
                self.sig_log.emit(
                    f"Auto Tune complete — CF={tune_data['center_freq_hz']/1e6:.4f} MHz  "
                    f"Span={tune_data['span_hz']/1e3:.1f} kHz  "
                    f"RBW={tune_data['rbw_hz']:.0f} Hz  "
                    f"VBW={tune_data['vbw_hz']:.0f} Hz  "
                    f"RefLev={tune_data['ref_level_dbm']:.1f} dBm"
                )

            # ---- Reference Gain (5 samples) -------------------------
            self.sig_log.emit("Measuring Reference Gain (5 samples)…")
            self.sig_status.emit("Calculating Reference Gain…")
            ref_gains = []
            for i in range(5):
                if self._stopped():
                    self._safe_rf_off()
                    return
                sg.set_power_dbm(p["start_power"])
                time.sleep(p["settle_time"])
                out_pwr = self._read_output_power(p)
                in_pwr  = p["start_power"] - p["in_loss"]
                ref_gains.append(out_pwr - in_pwr)

            ref_gain = sum(ref_gains) / len(ref_gains)
            self.sig_log.emit(f"Reference Gain = {ref_gain:.3f} dB")
            self.sig_measurement.emit({"ref_gain": ref_gain})

            # ---- Build sweep steps -----------------------------------
            steps = []
            pw = p["start_power"]
            while pw <= p["stop_power"] + 1e-9:
                steps.append(round(pw, 6))
                pw += p["step_size"]

            total = len(steps)
            self.sig_log.emit(f"Sweep: {total} steps from {p['start_power']} to {p['stop_power']} dBm")
            self.sig_status.emit("Sweep Running…")

            # ---- Sweep -----------------------------------------------
            for idx, set_pwr in enumerate(steps):
                if self._stopped():
                    self.sig_log.emit("Sweep stopped by user.")
                    break

                sg.set_power_dbm(set_pwr)
                time.sleep(p["settle_time"])

                # Peak Search
                if p.get("peak_search", True):
                    marker_freq, marker_amp = sa.peak_search()
                else:
                    settings = sa.get_settings()
                    marker_freq = settings["center_freq_hz"]
                    if p["meas_type"] == "Channel Power":
                        marker_amp = sa.read_channel_power()
                    else:
                        _, marker_amp = sa.peak_search()

                # Apply cable losses
                in_pwr  = set_pwr - p["in_loss"]
                out_pwr = marker_amp + p["out_loss"]
                gain    = out_pwr - in_pwr
                comp    = ref_gain - gain

                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                p1db_hit = comp >= 1.0

                # Log to Excel
                self.reporter.append_row(ts, in_pwr, out_pwr, gain, comp, highlight=p1db_hit)

                meas = {
                    "step": idx + 1, "total": total,
                    "set_power": set_pwr,
                    "in_pwr": in_pwr, "out_pwr": out_pwr,
                    "gain": gain, "compression": comp,
                    "marker_freq": marker_freq, "marker_amp": marker_amp,
                }
                self.sig_measurement.emit(meas)
                self.sig_progress.emit(idx + 1, total)
                self.sig_log.emit(
                    f"Step {idx+1}/{total}  In={in_pwr:.2f} dBm  "
                    f"Out={out_pwr:.2f} dBm  G={gain:.2f} dB  Comp={comp:.3f} dB"
                )

                if p1db_hit:
                    self.sig_log.emit(f"★ P1dB Reached at Input = {in_pwr:.2f} dBm")
                    self.sig_p1db_found.emit({
                        "in_pwr": in_pwr, "out_pwr": out_pwr,
                        "gain": gain, "compression": comp,
                        "step": idx + 1,
                    })
                    break

            # ---- RF OFF & Save ---------------------------------------
            self._safe_rf_off()
            self.reporter.save()
            self.sig_excel_saved.emit(str(self.reporter.filepath))
            self.sig_log.emit(f"Excel saved: {self.reporter.filepath}")
            self.sig_status.emit("Test Completed Successfully")
            self.sig_finished.emit()

        except Exception as exc:
            tb = traceback.format_exc()
            logger.error("TestEngine error:\n%s", tb)
            self._safe_rf_off()
            try:
                self.reporter.save()
            except Exception:
                pass
            self.sig_error.emit(str(exc))

    def _read_output_power(self, p: dict) -> float:
        if p["meas_type"] == "Channel Power":
            return self.sa.read_channel_power() + p["out_loss"]
        else:
            _, amp = self.sa.peak_search()
            return amp + p["out_loss"]

    def _safe_rf_off(self):
        try:
            self.sg.rf_off()
            self.sig_log.emit("RF OFF")
        except Exception as exc:
            logger.warning("RF OFF failed: %s", exc)


# ===========================================================================
# DISCOVERY WORKER  (QThread)
# ===========================================================================

class DiscoveryWorker(QThread):
    sig_log     = pyqtSignal(str)
    sig_done    = pyqtSignal(list, list)   # sg_addresses, sa_addresses
    sig_error   = pyqtSignal(str)

    def __init__(self, manager: InstrumentManager):
        super().__init__()
        self.manager = manager

    def run(self):
        try:
            sg_list, sa_list = self.manager.discover(progress_cb=self.sig_log.emit)
            self.sig_done.emit(sg_list, sa_list)
        except Exception as exc:
            self.sig_error.emit(str(exc))


# ===========================================================================
# THEME MANAGER
# ===========================================================================

class ThemeManager:
    DARK = """
        QMainWindow, QDialog { background:#1a1a2e; color:#e0e0e0; }
        QWidget { background:#1a1a2e; color:#e0e0e0; font-family:'Segoe UI', Arial; font-size:11px; }
        QGroupBox {
            border:1px solid #2d6a9f; border-radius:6px; margin-top:10px;
            padding-top:8px; font-weight:bold; color:#4fc3f7;
        }
        QGroupBox::title { subcontrol-origin:margin; left:10px; padding:0 4px; }
        QLineEdit, QDoubleSpinBox, QSpinBox, QComboBox {
            background:#0d1117; border:1px solid #2d6a9f; border-radius:4px;
            padding:4px 8px; color:#e0e0e0; selection-background-color:#2d6a9f;
        }
        QLineEdit:focus, QDoubleSpinBox:focus, QSpinBox:focus, QComboBox:focus {
            border:1px solid #4fc3f7;
        }
        QPushButton {
            background:#16213e; border:1px solid #2d6a9f; border-radius:5px;
            padding:6px 14px; color:#e0e0e0; font-weight:bold;
        }
        QPushButton:hover { background:#0f3460; border-color:#4fc3f7; }
        QPushButton:pressed { background:#2d6a9f; }
        QPushButton:disabled { background:#111; color:#555; border-color:#333; }
        QTextEdit {
            background:#0d1117; border:1px solid #2d6a9f; border-radius:4px;
            color:#b0bec5; font-family:'Consolas','Courier New',monospace; font-size:10px;
        }
        QProgressBar {
            border:1px solid #2d6a9f; border-radius:5px; background:#0d1117;
            text-align:center; color:#e0e0e0;
        }
        QProgressBar::chunk { background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #1565c0,stop:1 #4fc3f7); border-radius:4px; }
        QTabWidget::pane { border:1px solid #2d6a9f; border-radius:6px; }
        QTabBar::tab { background:#16213e; color:#aaa; padding:6px 16px; border-radius:4px 4px 0 0; }
        QTabBar::tab:selected { background:#0f3460; color:#4fc3f7; }
        QLabel#section_title { color:#4fc3f7; font-weight:bold; font-size:13px; }
        QStatusBar { background:#0d1117; color:#aaa; border-top:1px solid #2d6a9f; }
        QSplitter::handle { background:#2d6a9f; }
        QCheckBox { color:#e0e0e0; }
        QCheckBox::indicator { width:16px; height:16px; border:1px solid #2d6a9f; border-radius:3px; background:#0d1117; }
        QCheckBox::indicator:checked { background:#4fc3f7; border-color:#4fc3f7; }
        QScrollBar:vertical { background:#0d1117; width:10px; }
        QScrollBar::handle:vertical { background:#2d6a9f; border-radius:5px; }
        QComboBox QAbstractItemView { background:#0d1117; selection-background-color:#2d6a9f; color:#e0e0e0; }
    """

    LIGHT = """
        QMainWindow, QDialog { background:#f5f7fa; color:#1a1a2e; }
        QWidget { background:#f5f7fa; color:#1a1a2e; font-family:'Segoe UI',Arial; font-size:11px; }
        QGroupBox {
            border:1px solid #90caf9; border-radius:6px; margin-top:10px;
            padding-top:8px; font-weight:bold; color:#1565c0;
        }
        QGroupBox::title { subcontrol-origin:margin; left:10px; padding:0 4px; }
        QLineEdit, QDoubleSpinBox, QSpinBox, QComboBox {
            background:#ffffff; border:1px solid #90caf9; border-radius:4px;
            padding:4px 8px; color:#1a1a2e;
        }
        QLineEdit:focus, QDoubleSpinBox:focus, QSpinBox:focus, QComboBox:focus {
            border:1px solid #1565c0;
        }
        QPushButton {
            background:#e3f2fd; border:1px solid #90caf9; border-radius:5px;
            padding:6px 14px; color:#1565c0; font-weight:bold;
        }
        QPushButton:hover { background:#bbdefb; border-color:#1565c0; }
        QPushButton:pressed { background:#90caf9; }
        QPushButton:disabled { background:#eee; color:#aaa; border-color:#ccc; }
        QTextEdit {
            background:#ffffff; border:1px solid #90caf9; border-radius:4px;
            color:#37474f; font-family:'Consolas','Courier New',monospace; font-size:10px;
        }
        QProgressBar {
            border:1px solid #90caf9; border-radius:5px; background:#e3f2fd;
            text-align:center; color:#1a1a2e;
        }
        QProgressBar::chunk { background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #1565c0,stop:1 #42a5f5); border-radius:4px; }
        QTabWidget::pane { border:1px solid #90caf9; border-radius:6px; }
        QTabBar::tab { background:#e3f2fd; color:#555; padding:6px 16px; border-radius:4px 4px 0 0; }
        QTabBar::tab:selected { background:#bbdefb; color:#1565c0; }
        QLabel#section_title { color:#1565c0; font-weight:bold; font-size:13px; }
        QStatusBar { background:#e3f2fd; color:#555; border-top:1px solid #90caf9; }
        QCheckBox { color:#1a1a2e; }
        QCheckBox::indicator { width:16px; height:16px; border:1px solid #90caf9; border-radius:3px; background:#fff; }
        QCheckBox::indicator:checked { background:#1565c0; border-color:#1565c0; }
    """


# ===========================================================================
# INDICATOR WIDGET
# ===========================================================================

class ConnectionIndicator(QFrame):
    """Small colored circle indicator (Red / Green / Yellow)."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(18, 18)
        self._color = "red"
        self._update_style()

    def set_state(self, state: str):
        """state: 'red' | 'green' | 'yellow'"""
        self._color = state
        self._update_style()

    def _update_style(self):
        colors = {
            "red":    ("#e53935", "#b71c1c"),
            "green":  ("#43a047", "#1b5e20"),
            "yellow": ("#fdd835", "#f9a825"),
        }
        fg, dark = colors.get(self._color, ("#e53935", "#b71c1c"))
        self.setStyleSheet(
            f"border-radius:9px; background:qradialgradient("
            f"cx:0.35,cy:0.35,radius:0.65,fx:0.35,fy:0.35,"
            f"stop:0 {fg}, stop:1 {dark});"
        )


# ===========================================================================
# MAIN WINDOW
# ===========================================================================

class MainWindow(QMainWindow):
    """
    Primary application window.
    All UI creation is done in _build_ui().
    Thread-safe updates via Qt signals.
    """

    def __init__(self):
        super().__init__()
        self.setWindowTitle("P1dB Compression Point Test Automation  —  Keysight")
        self.resize(1440, 900)
        self.setMinimumSize(1100, 700)

        self._theme = "dark"
        self._settings = QSettings("P1dBApp", "P1dBTester")

        # State
        self._sg_inst:  Instrument | None = None
        self._sa_inst:  Instrument | None = None
        self._sg_scpi:  SignalGeneratorSCPI | None = None
        self._sa_scpi:  SignalAnalyzerSCPI  | None = None
        self._manager:  InstrumentManager | None = None
        self._engine:   TestEngine | None = None
        self._reporter: ExcelReporter | None = None
        self._disc_worker: DiscoveryWorker | None = None

        self._sg_addresses: list[str] = []
        self._sa_addresses: list[str] = []

        self._build_ui()
        self._apply_theme(self._theme)

    # ==================================================================
    # UI CONSTRUCTION
    # ==================================================================

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(8, 8, 8, 4)
        root.setSpacing(6)

        # Top toolbar row
        root.addLayout(self._build_toolbar())

        # Main content splitter
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setChildrenCollapsible(False)
        root.addWidget(splitter, stretch=1)

        # Left panel
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setFrameShape(QFrame.Shape.NoFrame)
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setSpacing(6)
        left_layout.setContentsMargins(0, 0, 4, 0)
        left_scroll.setWidget(left_widget)

        left_layout.addWidget(self._build_test_info())
        left_layout.addWidget(self._build_connection_panel())
        left_layout.addWidget(self._build_freq_config())
        left_layout.addWidget(self._build_analyzer_config())
        left_layout.addWidget(self._build_cable_loss())
        left_layout.addWidget(self._build_meas_config())
        left_layout.addWidget(self._build_sweep_config())
        left_layout.addStretch()

        splitter.addWidget(left_scroll)

        # Right panel (tabs)
        right_tabs = QTabWidget()
        right_tabs.addTab(self._build_measurements_tab(), "Live Measurements")
        right_tabs.addTab(self._build_activity_tab(), "Activity Log")
        splitter.addWidget(right_tabs)

        splitter.setSizes([480, 960])

        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self._set_status("Ready")

    # ---- Toolbar ---------------------------------------------------------
    def _build_toolbar(self) -> QHBoxLayout:
        layout = QHBoxLayout()
        layout.setSpacing(8)

        title = QLabel("P1dB Compression Point Test Automation")
        title.setObjectName("section_title")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        layout.addWidget(title)
        layout.addStretch()

        # Theme toggle
        self.btn_theme = QPushButton("☀ Light Theme")
        self.btn_theme.setFixedWidth(130)
        self.btn_theme.clicked.connect(self._toggle_theme)
        layout.addWidget(self.btn_theme)

        return layout

    # ---- Test Information -----------------------------------------------
    def _build_test_info(self) -> QGroupBox:
        grp = QGroupBox("Test Information")
        g = QGridLayout(grp)
        g.setSpacing(6)

        self.le_project   = QLineEdit(placeholderText="e.g. PA_Project")
        self.le_unit      = QLineEdit(placeholderText="e.g. Unit01")
        self.le_condition = QLineEdit(placeholderText="e.g. 25C")

        for row, (lbl, widget) in enumerate([
            ("Project Name",    self.le_project),
            ("Unit Name",       self.le_unit),
            ("Test Condition",  self.le_condition),
        ]):
            g.addWidget(QLabel(lbl), row, 0)
            g.addWidget(widget,     row, 1)

        return grp

    # ---- Connection Panel -----------------------------------------------
    def _build_connection_panel(self) -> QGroupBox:
        grp = QGroupBox("Instrument Connection")
        v = QVBoxLayout(grp)
        v.setSpacing(6)

        # Comm type
        h1 = QHBoxLayout()
        h1.addWidget(QLabel("Comm Type:"))
        self.cmb_conn_type = QComboBox()
        self.cmb_conn_type.addItems(["Auto Detect", "USB Only", "LAN Only"])
        h1.addWidget(self.cmb_conn_type)
        v.addLayout(h1)

        # Connect / Disconnect buttons
        h2 = QHBoxLayout()
        self.btn_connect    = QPushButton("Connect")
        self.btn_disconnect = QPushButton("Disconnect")
        self.btn_disconnect.setEnabled(False)
        self.btn_connect.clicked.connect(self._on_connect)
        self.btn_disconnect.clicked.connect(self._on_disconnect)
        h2.addWidget(self.btn_connect)
        h2.addWidget(self.btn_disconnect)
        v.addLayout(h2)

        # Instrument info grid
        g = QGridLayout()
        g.setSpacing(4)

        # SG row
        self.ind_sg = ConnectionIndicator()
        g.addWidget(self.ind_sg, 0, 0)
        g.addWidget(QLabel("Signal Generator"), 0, 1)
        self.lbl_sg_model  = QLabel("—")
        self.lbl_sg_serial = QLabel("—")
        self.lbl_sg_fw     = QLabel("—")
        self.lbl_sg_addr   = QLabel("—")
        self.lbl_sg_ctype  = QLabel("—")
        g.addWidget(self.lbl_sg_model,  0, 2)
        g.addWidget(QLabel("SN:"),      1, 1)
        g.addWidget(self.lbl_sg_serial, 1, 2)
        g.addWidget(QLabel("FW:"),      2, 1)
        g.addWidget(self.lbl_sg_fw,     2, 2)
        g.addWidget(QLabel("Addr:"),    3, 1)
        g.addWidget(self.lbl_sg_addr,   3, 2)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        g.addWidget(sep, 4, 0, 1, 3)

        # SA row
        self.ind_sa = ConnectionIndicator()
        g.addWidget(self.ind_sa, 5, 0)
        g.addWidget(QLabel("Signal Analyzer"), 5, 1)
        self.lbl_sa_model  = QLabel("—")
        self.lbl_sa_serial = QLabel("—")
        self.lbl_sa_fw     = QLabel("—")
        self.lbl_sa_addr   = QLabel("—")
        g.addWidget(self.lbl_sa_model,  5, 2)
        g.addWidget(QLabel("SN:"),      6, 1)
        g.addWidget(self.lbl_sa_serial, 6, 2)
        g.addWidget(QLabel("FW:"),      7, 1)
        g.addWidget(self.lbl_sa_fw,     7, 2)
        g.addWidget(QLabel("Addr:"),    8, 1)
        g.addWidget(self.lbl_sa_addr,   8, 2)

        v.addLayout(g)
        return grp

    # ---- Frequency Configuration ----------------------------------------
    def _build_freq_config(self) -> QGroupBox:
        grp = QGroupBox("Frequency Configuration")
        g = QGridLayout(grp)
        g.setSpacing(6)

        self.dsb_freq = QDoubleSpinBox()
        self.dsb_freq.setRange(0.001, 999999.999)
        self.dsb_freq.setDecimals(6)
        self.dsb_freq.setValue(2450.0)
        self.dsb_freq.setStepType(QDoubleSpinBox.StepType.AdaptiveDecimalStepType)

        self.cmb_freq_unit = QComboBox()
        self.cmb_freq_unit.addItems(["Hz", "kHz", "MHz", "GHz"])
        self.cmb_freq_unit.setCurrentText("MHz")

        g.addWidget(QLabel("Center Frequency:"), 0, 0)
        g.addWidget(self.dsb_freq,               0, 1)
        g.addWidget(self.cmb_freq_unit,          0, 2)

        return grp

    # ---- Analyzer Config ------------------------------------------------
    def _build_analyzer_config(self) -> QGroupBox:
        grp = QGroupBox("Analyzer Configuration")
        g = QGridLayout(grp)
        g.setSpacing(6)

        self.dsb_ref_level = QDoubleSpinBox()
        self.dsb_ref_level.setRange(-200.0, 0.0)
        self.dsb_ref_level.setValue(-10.0)
        self.dsb_ref_level.setSuffix(" dBm")

        g.addWidget(QLabel("Reference Level (≤ 0 dBm):"), 0, 0)
        g.addWidget(self.dsb_ref_level, 0, 1)

        return grp

    # ---- Cable Loss -----------------------------------------------------
    def _build_cable_loss(self) -> QGroupBox:
        grp = QGroupBox("Cable Loss Configuration")
        g = QGridLayout(grp)
        g.setSpacing(6)

        self.dsb_in_loss  = QDoubleSpinBox()
        self.dsb_out_loss = QDoubleSpinBox()
        for sp in (self.dsb_in_loss, self.dsb_out_loss):
            sp.setRange(0.0, 50.0)
            sp.setDecimals(2)
            sp.setSuffix(" dB")

        g.addWidget(QLabel("Input Cable Loss:"),  0, 0)
        g.addWidget(self.dsb_in_loss,             0, 1)
        g.addWidget(QLabel("Output Cable Loss:"), 1, 0)
        g.addWidget(self.dsb_out_loss,            1, 1)

        return grp

    # ---- Measurement Config --------------------------------------------
    def _build_meas_config(self) -> QGroupBox:
        grp = QGroupBox("Measurement Configuration")
        v = QVBoxLayout(grp)
        v.setSpacing(6)

        h1 = QHBoxLayout()
        h1.addWidget(QLabel("Measurement Type:"))
        self.cmb_meas_type = QComboBox()
        self.cmb_meas_type.addItems(["Marker Power", "Channel Power"])
        self.cmb_meas_type.currentTextChanged.connect(self._on_meas_type_changed)
        h1.addWidget(self.cmb_meas_type)
        v.addLayout(h1)

        # Channel BW (visible only for Channel Power)
        self.ch_bw_widget = QWidget()
        h2 = QHBoxLayout(self.ch_bw_widget)
        h2.setContentsMargins(0,0,0,0)
        h2.addWidget(QLabel("Channel BW:"))
        self.dsb_ch_bw = QDoubleSpinBox()
        self.dsb_ch_bw.setRange(0.001, 1000.0)
        self.dsb_ch_bw.setValue(10.0)
        h2.addWidget(self.dsb_ch_bw)
        self.cmb_ch_bw_unit = QComboBox()
        self.cmb_ch_bw_unit.addItems(["Hz", "kHz", "MHz"])
        self.cmb_ch_bw_unit.setCurrentText("MHz")
        h2.addWidget(self.cmb_ch_bw_unit)
        self.ch_bw_widget.setVisible(False)
        v.addWidget(self.ch_bw_widget)

        self.chk_auto_tune   = QCheckBox("Auto Tune Before Measurement")
        self.chk_peak_search = QCheckBox("Auto Peak Search During Sweep")
        self.chk_auto_tune.setChecked(True)
        self.chk_peak_search.setChecked(True)
        v.addWidget(self.chk_auto_tune)
        v.addWidget(self.chk_peak_search)

        return grp

    # ---- Sweep Config ---------------------------------------------------
    def _build_sweep_config(self) -> QGroupBox:
        grp = QGroupBox("Power Sweep Configuration")
        g = QGridLayout(grp)
        g.setSpacing(6)

        self.dsb_start = QDoubleSpinBox()
        self.dsb_stop  = QDoubleSpinBox()
        self.dsb_step  = QDoubleSpinBox()
        self.dsb_settle = QDoubleSpinBox()

        self.dsb_start.setRange(-120.0, 30.0); self.dsb_start.setValue(-30.0); self.dsb_start.setSuffix(" dBm")
        self.dsb_stop.setRange (-120.0, 30.0); self.dsb_stop.setValue( 20.0); self.dsb_stop.setSuffix(" dBm")
        self.dsb_step.setRange (0.1, 10.0);    self.dsb_step.setValue(  1.0); self.dsb_step.setSuffix(" dB");  self.dsb_step.setDecimals(2)
        self.dsb_settle.setRange(0.05, 5.0);   self.dsb_settle.setValue(0.3); self.dsb_settle.setSuffix(" s");  self.dsb_settle.setDecimals(2)

        rows = [
            ("Start Power:", self.dsb_start),
            ("Stop Power:",  self.dsb_stop),
            ("Step Size:",   self.dsb_step),
            ("Settle Time:", self.dsb_settle),
        ]
        for r, (lbl, wgt) in enumerate(rows):
            g.addWidget(QLabel(lbl), r, 0)
            g.addWidget(wgt,         r, 1)

        # RF & Test controls
        h = QHBoxLayout()
        self.btn_rf_on  = QPushButton("RF ON")
        self.btn_rf_off = QPushButton("RF OFF")
        self.btn_start  = QPushButton("▶  Start Test")
        self.btn_stop   = QPushButton("■  Stop Test")

        self.btn_rf_on.setStyleSheet("background:#1b5e20; color:white; font-weight:bold;")
        self.btn_rf_off.setStyleSheet("background:#b71c1c; color:white; font-weight:bold;")
        self.btn_start.setStyleSheet("background:#0d47a1; color:white; font-weight:bold;")
        self.btn_stop.setStyleSheet("background:#6d1b7b; color:white; font-weight:bold;")

        self.btn_rf_off.setEnabled(False)
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(False)

        self.btn_rf_on.clicked.connect(self._on_rf_on)
        self.btn_rf_off.clicked.connect(self._on_rf_off)
        self.btn_start.clicked.connect(self._on_start_test)
        self.btn_stop.clicked.connect(self._on_stop_test)

        for btn in (self.btn_rf_on, self.btn_rf_off, self.btn_start, self.btn_stop):
            h.addWidget(btn)

        g.addLayout(h, len(rows), 0, 1, 2)
        return grp

    # ---- Measurements Tab -----------------------------------------------
    def _build_measurements_tab(self) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.setSpacing(8)

        # Auto Tune results
        tune_grp = QGroupBox("Auto Tune Results")
        tg = QGridLayout(tune_grp)
        self._tune_labels = {}
        for i, (key, label) in enumerate([
            ("center_freq", "Center Frequency"),
            ("span",        "Span"),
            ("rbw",         "RBW"),
            ("vbw",         "VBW"),
            ("ref_level",   "Reference Level"),
        ]):
            tg.addWidget(QLabel(label + ":"), i, 0)
            lbl = QLabel("—")
            lbl.setFont(QFont("Consolas", 11))
            self._tune_labels[key] = lbl
            tg.addWidget(lbl, i, 1)
        v.addWidget(tune_grp)

        # Real-time readings
        live_grp = QGroupBox("Real-Time Measurements")
        lg = QGridLayout(live_grp)
        self._live_labels = {}
        items = [
            ("ref_gain",    "Reference Gain (dB)"),
            ("in_pwr",      "Input Power (dBm)"),
            ("out_pwr",     "Output Power (dBm)"),
            ("gain",        "Gain (dB)"),
            ("compression", "Compression (dB)"),
            ("marker_freq", "Marker Frequency (MHz)"),
            ("marker_amp",  "Marker Amplitude (dBm)"),
        ]
        for i, (key, label) in enumerate(items):
            lg.addWidget(QLabel(label + ":"), i, 0)
            lbl = QLabel("—")
            lbl.setFont(QFont("Consolas", 13, QFont.Weight.Bold))
            self._live_labels[key] = lbl
            lg.addWidget(lbl, i, 1)
        v.addWidget(live_grp)

        # Progress
        prog_grp = QGroupBox("Sweep Progress")
        pg = QVBoxLayout(prog_grp)
        self.lbl_step = QLabel("Step — / —")
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        pg.addWidget(self.lbl_step)
        pg.addWidget(self.progress_bar)
        v.addWidget(prog_grp)

        # P1dB result
        self.p1db_grp = QGroupBox("★  P1dB Result")
        p1g = QGridLayout(self.p1db_grp)
        self._p1db_labels = {}
        for i, (key, lbl) in enumerate([
            ("in_pwr",     "Input Power at P1dB"),
            ("out_pwr",    "Output Power at P1dB"),
            ("gain",       "Gain at P1dB"),
            ("compression","Compression"),
        ]):
            p1g.addWidget(QLabel(lbl + ":"), i, 0)
            l = QLabel("—")
            l.setFont(QFont("Consolas", 13, QFont.Weight.Bold))
            l.setStyleSheet("color:#43a047;")
            self._p1db_labels[key] = l
            p1g.addWidget(l, i, 1)
        self.p1db_grp.setVisible(False)
        v.addWidget(self.p1db_grp)

        v.addStretch()
        return w

    # ---- Activity Tab ---------------------------------------------------
    def _build_activity_tab(self) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        self.log_view = QTextEdit()
        self.log_view.setReadOnly(True)
        v.addWidget(self.log_view)
        h = QHBoxLayout()
        btn_clear = QPushButton("Clear Log")
        btn_clear.clicked.connect(self.log_view.clear)
        btn_save_log = QPushButton("Save Log…")
        btn_save_log.clicked.connect(self._save_log)
        h.addStretch()
        h.addWidget(btn_clear)
        h.addWidget(btn_save_log)
        v.addLayout(h)
        return w

    # ==================================================================
    # THEME
    # ==================================================================

    def _apply_theme(self, theme: str):
        self._theme = theme
        if theme == "dark":
            self.setStyleSheet(ThemeManager.DARK)
            self.btn_theme.setText("☀ Light Theme")
        else:
            self.setStyleSheet(ThemeManager.LIGHT)
            self.btn_theme.setText("🌙 Dark Theme")

    def _toggle_theme(self):
        self._apply_theme("light" if self._theme == "dark" else "dark")

    # ==================================================================
    # HELPERS
    # ==================================================================

    def _freq_hz(self) -> float:
        val  = self.dsb_freq.value()
        unit = self.cmb_freq_unit.currentText()
        mult = {"Hz": 1, "kHz": 1e3, "MHz": 1e6, "GHz": 1e9}[unit]
        return val * mult

    def _ch_bw_hz(self) -> float:
        val  = self.dsb_ch_bw.value()
        unit = self.cmb_ch_bw_unit.currentText()
        mult = {"Hz": 1, "kHz": 1e3, "MHz": 1e6}[unit]
        return val * mult

    def _set_status(self, msg: str):
        self.status_bar.showMessage(msg)

    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        line = f"[{ts}]  {msg}"
        self.log_view.append(line)
        # Auto scroll
        cursor = self.log_view.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        self.log_view.setTextCursor(cursor)
        logger.info(msg)

    def _validate_inputs(self) -> bool:
        if self.dsb_ref_level.value() > 0:
            QMessageBox.warning(self, "Validation Error", "Reference Level must be ≤ 0 dBm.")
            return False
        if self.dsb_start.value() >= self.dsb_stop.value():
            QMessageBox.warning(self, "Validation Error", "Start Power must be less than Stop Power.")
            return False
        if self.le_project.text().strip() == "":
            QMessageBox.warning(self, "Validation Error", "Project Name is required.")
            return False
        return True

    def _conn_type_filter(self) -> str:
        t = self.cmb_conn_type.currentText()
        if "USB" in t:
            return "USB"
        if "LAN" in t:
            return "LAN"
        return "AUTO"

    # ==================================================================
    # SLOT: Measurement type changed
    # ==================================================================

    def _on_meas_type_changed(self, text: str):
        self.ch_bw_widget.setVisible(text == "Channel Power")

    # ==================================================================
    # SLOT: Connect
    # ==================================================================

    @pyqtSlot()
    def _on_connect(self):
        if pyvisa is None:
            QMessageBox.critical(self, "Error",
                "PyVISA is not installed.\n\npip install pyvisa pyvisa-py")
            return

        self.btn_connect.setEnabled(False)
        self._set_status("Searching for instruments…")
        self._log("Starting instrument discovery…")

        self._manager = InstrumentManager(filter_mode=self._conn_type_filter())
        self._disc_worker = DiscoveryWorker(self._manager)
        self._disc_worker.sig_log.connect(self._log)
        self._disc_worker.sig_done.connect(self._on_discovery_done)
        self._disc_worker.sig_error.connect(self._on_discovery_error)
        self._disc_worker.start()

    @pyqtSlot(list, list)
    def _on_discovery_done(self, sg_list: list, sa_list: list):
        self._sg_addresses = sg_list
        self._sa_addresses = sa_list

        if not sg_list:
            self._log("⚠ No Signal Generator found.")
            QMessageBox.warning(self, "No Signal Generator",
                "No compatible Keysight Signal Generator was detected.\n"
                "Check connections and communication type.")
            self.btn_connect.setEnabled(True)
            return

        if not sa_list:
            self._log("⚠ No Signal Analyzer found.")
            QMessageBox.warning(self, "No Signal Analyzer",
                "No compatible Keysight Signal Analyzer was detected.")
            self.btn_connect.setEnabled(True)
            return

        # Connect first found of each type
        try:
            sg = self._manager.connect_sg(sg_list[0])
            self._sg_inst = sg
            self._sg_scpi = SignalGeneratorSCPI(sg)
            self.ind_sg.set_state("green")
            self.lbl_sg_model.setText(f"{sg.manufacturer} {sg.model}")
            self.lbl_sg_serial.setText(sg.serial)
            self.lbl_sg_fw.setText(sg.firmware)
            self.lbl_sg_addr.setText(sg.address)
            self._log(f"✔ SG Connected: {sg.model}  SN:{sg.serial}")
        except Exception as exc:
            self._log(f"✘ SG Connection failed: {exc}")
            QMessageBox.critical(self, "SG Error", str(exc))
            self.btn_connect.setEnabled(True)
            return

        try:
            sa = self._manager.connect_sa(sa_list[0])
            self._sa_inst = sa
            self._sa_scpi = SignalAnalyzerSCPI(sa)
            self.ind_sa.set_state("green")
            self.lbl_sa_model.setText(f"{sa.manufacturer} {sa.model}")
            self.lbl_sa_serial.setText(sa.serial)
            self.lbl_sa_fw.setText(sa.firmware)
            self.lbl_sa_addr.setText(sa.address)
            self._log(f"✔ SA Connected: {sa.model}  SN:{sa.serial}")
        except Exception as exc:
            self._log(f"✘ SA Connection failed: {exc}")
            QMessageBox.critical(self, "SA Error", str(exc))
            self.btn_connect.setEnabled(True)
            return

        # Update UI state
        self.btn_connect.setText("CONNECTED")
        self.btn_connect.setStyleSheet("background:#1b5e20; color:white; font-weight:bold;")
        self.btn_connect.setEnabled(False)
        self.btn_disconnect.setEnabled(True)
        self.btn_rf_on.setEnabled(True)
        self.btn_start.setEnabled(True)
        self._set_status("Instruments Connected")

    @pyqtSlot(str)
    def _on_discovery_error(self, msg: str):
        self._log(f"✘ Discovery error: {msg}")
        QMessageBox.critical(self, "Discovery Error", msg)
        self.btn_connect.setEnabled(True)
        self._set_status("Error — see log")

    # ==================================================================
    # SLOT: Disconnect
    # ==================================================================

    @pyqtSlot()
    def _on_disconnect(self):
        self._safe_rf_off()
        if self._manager:
            self._manager.disconnect_all()
        self._sg_inst = self._sa_inst = None
        self._sg_scpi = self._sa_scpi = None
        self.ind_sg.set_state("red")
        self.ind_sa.set_state("red")
        for lbl in (self.lbl_sg_model, self.lbl_sg_serial, self.lbl_sg_fw,
                    self.lbl_sg_addr, self.lbl_sa_model, self.lbl_sa_serial,
                    self.lbl_sa_fw,   self.lbl_sa_addr):
            lbl.setText("—")
        self.btn_connect.setText("Connect")
        self.btn_connect.setStyleSheet("")
        self.btn_connect.setEnabled(True)
        self.btn_disconnect.setEnabled(False)
        self.btn_rf_on.setEnabled(False)
        self.btn_rf_off.setEnabled(False)
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(False)
        self._log("All instruments disconnected.")
        self._set_status("Disconnected")

    # ==================================================================
    # SLOT: RF ON / OFF
    # ==================================================================

    @pyqtSlot()
    def _on_rf_on(self):
        if not self._sg_scpi:
            return
        try:
            self._sg_scpi.set_frequency_hz(self._freq_hz())
            self._sg_scpi.set_power_dbm(self.dsb_start.value())
            self._sg_scpi.rf_on()
            self.btn_rf_on.setEnabled(False)
            self.btn_rf_off.setEnabled(True)
            self._log("RF ON")
            self._set_status("RF ON")

            if self.chk_auto_tune.isChecked() and self._sa_scpi:
                self._log("Auto Tune executing…")
                tune_data = self._sa_scpi.auto_tune()
                self._update_tune_display(tune_data)
                self._log("Auto Tune complete.")

            if self.chk_peak_search.isChecked() and self._sa_scpi:
                freq, amp = self._sa_scpi.peak_search()
                self._live_labels["marker_freq"].setText(f"{freq/1e6:.6f} MHz")
                self._live_labels["marker_amp"].setText(f"{amp:.3f} dBm")
                self._log(f"Peak Search: {freq/1e6:.4f} MHz  {amp:.3f} dBm")

        except Exception as exc:
            self._log(f"RF ON error: {exc}")
            QMessageBox.critical(self, "RF Error", str(exc))

    @pyqtSlot()
    def _on_rf_off(self):
        self._safe_rf_off()
        self.btn_rf_on.setEnabled(True)
        self.btn_rf_off.setEnabled(False)
        self._set_status("RF OFF")

    def _safe_rf_off(self):
        if self._sg_scpi:
            try:
                self._sg_scpi.rf_off()
                self._log("RF OFF")
            except Exception as exc:
                self._log(f"RF OFF warning: {exc}")

    # ==================================================================
    # SLOT: Start / Stop Test
    # ==================================================================

    @pyqtSlot()
    def _on_start_test(self):
        if not self._validate_inputs():
            return
        if not (self._sg_scpi and self._sa_scpi):
            QMessageBox.warning(self, "Not Connected", "Connect instruments first.")
            return

        self._log("=== Test Started ===")
        self.p1db_grp.setVisible(False)
        self.progress_bar.setValue(0)

        # Build metadata dict for reporter
        meta = {
            "project_name":    self.le_project.text().strip() or "Project",
            "unit_name":       self.le_unit.text().strip() or "Unit",
            "test_condition":  self.le_condition.text().strip() or "Cond",
            "sg_manufacturer": self._sg_inst.manufacturer if self._sg_inst else "",
            "sg_model":        self._sg_inst.model        if self._sg_inst else "",
            "sg_serial":       self._sg_inst.serial       if self._sg_inst else "",
            "sa_manufacturer": self._sa_inst.manufacturer if self._sa_inst else "",
            "sa_model":        self._sa_inst.model        if self._sa_inst else "",
            "sa_serial":       self._sa_inst.serial       if self._sa_inst else "",
            "center_freq_str": f"{self.dsb_freq.value()} {self.cmb_freq_unit.currentText()}",
            "freq_unit":       self.cmb_freq_unit.currentText(),
            "ref_level":       self.dsb_ref_level.value(),
            "in_loss":         self.dsb_in_loss.value(),
            "out_loss":        self.dsb_out_loss.value(),
            "meas_type":       self.cmb_meas_type.currentText(),
            "conn_type":       self.cmb_conn_type.currentText(),
            "span_str":        "—", "rbw_str": "—", "vbw_str": "—",
            "marker_freq_str": "—",
        }

        self._reporter = ExcelReporter()
        try:
            path = self._reporter.create(meta)
            self._log(f"Excel file created: {path}")
        except Exception as exc:
            QMessageBox.critical(self, "Excel Error", f"Cannot create report:\n{exc}")
            return

        # Sweep parameters
        params = {
            "center_freq_hz": self._freq_hz(),
            "ref_level":      self.dsb_ref_level.value(),
            "in_loss":        self.dsb_in_loss.value(),
            "out_loss":       self.dsb_out_loss.value(),
            "meas_type":      self.cmb_meas_type.currentText(),
            "ch_bw_hz":       self._ch_bw_hz(),
            "start_power":    self.dsb_start.value(),
            "stop_power":     self.dsb_stop.value(),
            "step_size":      self.dsb_step.value(),
            "settle_time":    self.dsb_settle.value(),
            "auto_tune":      self.chk_auto_tune.isChecked(),
            "peak_search":    self.chk_peak_search.isChecked(),
        }

        self._engine = TestEngine(self._sg_scpi, self._sa_scpi, params, self._reporter)
        self._engine.sig_log.connect(self._log)
        self._engine.sig_status.connect(self._set_status)
        self._engine.sig_progress.connect(self._update_progress)
        self._engine.sig_measurement.connect(self._update_measurements)
        self._engine.sig_p1db_found.connect(self._on_p1db_found)
        self._engine.sig_error.connect(self._on_engine_error)
        self._engine.sig_finished.connect(self._on_test_finished)
        self._engine.sig_tune_data.connect(self._update_tune_display)
        self._engine.sig_excel_saved.connect(lambda p: self._log(f"Excel saved: {p}"))

        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.btn_rf_on.setEnabled(False)
        self.btn_rf_off.setEnabled(False)
        self._engine.start()
        self._set_status("Test Running…")

    @pyqtSlot()
    def _on_stop_test(self):
        if self._engine and self._engine.isRunning():
            self._engine.stop()
            self._log("Stop requested by user.")
            self._set_status("Stopping…")

    # ==================================================================
    # Engine callbacks
    # ==================================================================

    @pyqtSlot(int, int)
    def _update_progress(self, current: int, total: int):
        pct = int(current / total * 100) if total else 0
        self.progress_bar.setValue(pct)
        self.lbl_step.setText(f"Step {current} / {total}")

    @pyqtSlot(dict)
    def _update_measurements(self, d: dict):
        fmt = lambda v, dec=3: f"{v:.{dec}f}" if isinstance(v, float) else str(v)
        if "ref_gain" in d:
            self._live_labels["ref_gain"].setText(fmt(d["ref_gain"]) + " dB")
        if "in_pwr" in d:
            self._live_labels["in_pwr"].setText(fmt(d["in_pwr"]) + " dBm")
            self._live_labels["out_pwr"].setText(fmt(d["out_pwr"]) + " dBm")
            self._live_labels["gain"].setText(fmt(d["gain"]) + " dB")
            self._live_labels["compression"].setText(fmt(d["compression"]) + " dB")
            self._live_labels["marker_freq"].setText(fmt(d["marker_freq"] / 1e6, 4) + " MHz")
            self._live_labels["marker_amp"].setText(fmt(d["marker_amp"]) + " dBm")

    @pyqtSlot(dict)
    def _update_tune_display(self, d: dict):
        def _fmt_hz(v):
            if v >= 1e9:  return f"{v/1e9:.6f} GHz"
            if v >= 1e6:  return f"{v/1e6:.6f} MHz"
            if v >= 1e3:  return f"{v/1e3:.3f} kHz"
            return f"{v:.1f} Hz"

        self._tune_labels["center_freq"].setText(_fmt_hz(d.get("center_freq_hz", 0)))
        self._tune_labels["span"].setText(_fmt_hz(d.get("span_hz", 0)))
        self._tune_labels["rbw"].setText(_fmt_hz(d.get("rbw_hz", 0)))
        self._tune_labels["vbw"].setText(_fmt_hz(d.get("vbw_hz", 0)))
        self._tune_labels["ref_level"].setText(f"{d.get('ref_level_dbm', 0):.2f} dBm")

    @pyqtSlot(dict)
    def _on_p1db_found(self, d: dict):
        self.p1db_grp.setVisible(True)
        self._p1db_labels["in_pwr"].setText(f"{d['in_pwr']:.3f} dBm")
        self._p1db_labels["out_pwr"].setText(f"{d['out_pwr']:.3f} dBm")
        self._p1db_labels["gain"].setText(f"{d['gain']:.3f} dB")
        self._p1db_labels["compression"].setText(f"{d['compression']:.3f} dB")

        QMessageBox.information(
            self, "★ P1dB Compression Point Reached",
            f"P1dB Compression Point Detected!\n\n"
            f"  Input Power  : {d['in_pwr']:.3f} dBm\n"
            f"  Output Power : {d['out_pwr']:.3f} dBm\n"
            f"  Gain         : {d['gain']:.3f} dB\n"
            f"  Compression  : {d['compression']:.3f} dB\n\n"
            f"Test Completed Successfully."
        )

    @pyqtSlot(str)
    def _on_engine_error(self, msg: str):
        self._log(f"✘ ENGINE ERROR: {msg}")
        QMessageBox.critical(self, "Test Error", msg)
        self._reset_controls()

    @pyqtSlot()
    def _on_test_finished(self):
        self._log("=== Test Finished ===")
        self._reset_controls()

    def _reset_controls(self):
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.btn_rf_on.setEnabled(True)
        self.btn_rf_off.setEnabled(False)

    # ==================================================================
    # Save log
    # ==================================================================

    def _save_log(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Log", str(LOG_DIR / "activity_log.txt"),
            "Text Files (*.txt);;All Files (*)"
        )
        if path:
            with open(path, "w", encoding="utf-8") as f:
                f.write(self.log_view.toPlainText())
            self._log(f"Log saved: {path}")

    # ==================================================================
    # Close event
    # ==================================================================

    def closeEvent(self, event):
        reply = QMessageBox.question(
            self, "Confirm Exit",
            "Exit application?\nAll instruments will be disconnected safely.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            if self._engine and self._engine.isRunning():
                self._engine.stop()
                self._engine.wait(3000)
            self._safe_rf_off()
            if self._manager:
                self._manager.disconnect_all()
            logger.info("Application closed. All instruments released.")
            event.accept()
        else:
            event.ignore()


# ===========================================================================
# ENTRY POINT
# ===========================================================================

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("P1dB Tester")
    app.setOrganizationName("RF Test Automation")
    app.setStyle("Fusion")

    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()