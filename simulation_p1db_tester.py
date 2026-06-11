"""
P1dB Compression Point Test Automation Software - SIMULATION MODE (FIXED)
=================================================================
Complete working version - Fixed the boolean/callable conflict
"""

import sys
import os
import time
import logging
import traceback
import random
import math
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# SIMULATION MODE - Set to True for testing without instruments
# ---------------------------------------------------------------------------
USE_SIMULATION = True  # Keep as True for testing

# ---------------------------------------------------------------------------
# Third-Party Imports
# ---------------------------------------------------------------------------
try:
    import pyvisa
except ImportError:
    pyvisa = None
    print("Note: pyvisa not installed - simulation mode works without it")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
    print("Please install openpyxl: pip install openpyxl")

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QGroupBox, QLabel, QLineEdit, QPushButton, QComboBox,
    QCheckBox, QProgressBar, QTextEdit, QStatusBar, QFrame, QSplitter,
    QScrollArea, QMessageBox, QFileDialog, QTabWidget, QDoubleSpinBox,
    QSpinBox
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, pyqtSlot, QMutex, QMutexLocker,
    QSettings
)
from PyQt6.QtGui import QFont, QTextCursor

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------
LOG_DIR = Path("./P1dB_Test_Results/logs")
LOG_DIR.mkdir(parents=True, exist_ok=True)
_log_file = LOG_DIR / f"p1db_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(_log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("P1dB")


# ===========================================================================
# SIMULATION CLASSES
# ===========================================================================

class SimulatedInstrument:
    """Simulates a VISA instrument for testing."""
    
    def __init__(self, model="Simulated Keysight Instrument"):
        self.model = model
        self.serial = f"SIM-{random.randint(10000, 99999)}"
        self.firmware = "1.0.0"
        self.manufacturer = "Keysight Technologies"
        self.address = "SIMULATED::USB::TEST"
        
        # Simulation state
        self._frequency = 2.45e9
        self._power = -30.0
        self._rf_on = False
        self._ref_level = -10.0
        
    def query(self, cmd, timeout_ms=10000):
        """Simulate query response."""
        time.sleep(0.02)
        
        if "*IDN?" in cmd:
            return f"{self.manufacturer},{self.model},{self.serial},{self.firmware}"
        elif ":FREQ?" in cmd:
            return f"{self._frequency:.6f}"
        elif ":POW?" in cmd:
            return f"{self._power:.3f}"
        elif "CENT?" in cmd:
            return f"{self._frequency:.6f}"
        elif "RLEV?" in cmd:
            return f"{self._ref_level:.2f}"
        else:
            return "0"
    
    def write(self, cmd):
        """Simulate write command."""
        time.sleep(0.01)
        
        if "FREQ" in cmd and "Hz" in cmd:
            import re
            match = re.search(r"([\d\.]+)Hz", cmd)
            if match:
                self._frequency = float(match.group(1))
        elif "POW" in cmd and "dBm" in cmd:
            import re
            match = re.search(r"([\-\d\.]+)dBm", cmd)
            if match:
                self._power = float(match.group(1))
        elif "OUTP ON" in cmd:
            self._rf_on = True
        elif "OUTP OFF" in cmd:
            self._rf_on = False
    
    def close(self):
        pass


class SimulatedAmplifier:
    """Realistic amplifier model for simulation."""
    
    @staticmethod
    def calculate_output(input_power_dbm, p1db_input=-5.0, small_signal_gain=20.0):
        """
        Calculate output power with realistic compression.
        
        Args:
            input_power_dbm: Input power in dBm
            p1db_input: Input power at 1dB compression point (default -5 dBm)
            small_signal_gain: Small signal gain in dB (default 20 dB)
        
        Returns:
            output_power_dbm: Output power with compression
            gain: Actual gain in dB
            compression: Gain compression in dB
        """
        # Convert to linear for compression calculation
        input_linear = 10 ** ((input_power_dbm - 30) / 10)
        p1db_input_linear = 10 ** ((p1db_input - 30) / 10)
        
        # Rapp model for amplifier compression
        n = 2.5  # Smoothness factor
        vsat = p1db_input_linear * (10 ** (0.05))  # Saturation point ~1.26x P1dB
        
        if input_linear < 1e-12:
            output_linear = 0
        else:
            output_linear = input_linear / ((1 + (input_linear / vsat) ** (2 * n)) ** (1 / (2 * n)))
        
        # Add small signal gain
        gain_linear = 10 ** (small_signal_gain / 10)
        output_linear *= gain_linear
        
        # Convert back to dBm
        if output_linear > 0:
            output_power_dbm = 10 * math.log10(output_linear) + 30
        else:
            output_power_dbm = -100
        
        # Calculate actual gain and compression
        actual_gain = output_power_dbm - input_power_dbm
        compression = small_signal_gain - actual_gain
        
        return output_power_dbm, actual_gain, max(0, compression)


class SimulatedSignalGenerator:
    """Simulated signal generator."""
    
    def __init__(self):
        self.frequency = 2.45e9
        self.power = -30.0
        self.rf_enabled = False
        self.inst = SimulatedInstrument("Simulated MXG N5182B")
    
    def set_frequency_hz(self, freq_hz):
        self.frequency = freq_hz
        
    def set_power_dbm(self, power_dbm):
        self.power = power_dbm
        
    def rf_on(self):
        self.rf_enabled = True
        
    def rf_off(self):
        self.rf_enabled = False
        
    def get_power_dbm(self):
        return self.power


class SimulatedSignalAnalyzer:
    """Simulated signal analyzer with realistic amplifier model."""
    
    def __init__(self):
        self.center_freq = 2.45e9
        self.ref_level = -10.0
        self.inst = SimulatedInstrument("Simulated MXA N9020B")
        self._input_power = -30.0
        self.small_signal_gain = 20.0
        self.p1db_input = -5.0
        
    def set_frequency_hz(self, freq_hz):
        self.center_freq = freq_hz
        
    def set_reference_level(self, ref_dbm):
        self.ref_level = ref_dbm
        
    def set_input_power(self, power_dbm):
        """Track input power for simulation."""
        self._input_power = power_dbm
        
    def auto_tune(self):
        """Simulate auto-tune."""
        return {
            "center_freq_hz": self.center_freq,
            "span_hz": 10e6,
            "rbw_hz": 100e3,
            "vbw_hz": 100e3,
            "ref_level_dbm": self.ref_level
        }
        
    def peak_search(self):
        """Simulate peak search with realistic amplifier compression."""
        output_power, gain, compression = SimulatedAmplifier.calculate_output(
            self._input_power, 
            p1db_input=self.p1db_input,
            small_signal_gain=self.small_signal_gain
        )
        
        # Add small noise for realism
        output_power += random.uniform(-0.03, 0.03)
        
        return self.center_freq, output_power
    
    def read_channel_power(self):
        """Simulate channel power measurement."""
        _, power = self.peak_search()
        return power
    
    def get_settings(self):
        return {
            "center_freq_hz": self.center_freq,
            "span_hz": 10e6,
            "rbw_hz": 100e3,
            "vbw_hz": 100e3,
            "ref_level_dbm": self.ref_level,
        }


# ===========================================================================
# INSTRUMENT MANAGER (Simulation Mode)
# ===========================================================================

class InstrumentError(Exception):
    pass


class Instrument:
    def __init__(self, resource, address: str):
        self.resource = resource
        self.address = address
        self.manufacturer = ""
        self.model = ""
        self.serial = ""
        self.firmware = ""
        self.conn_type = ""

    def query(self, cmd: str, timeout_ms: int = 10_000) -> str:
        return self.resource.query(cmd)

    def write(self, cmd: str) -> None:
        self.resource.write(cmd)

    def identify(self) -> None:
        idn = self.query("*IDN?")
        parts = [p.strip() for p in idn.split(",")]
        self.manufacturer = parts[0] if len(parts) > 0 else "Unknown"
        self.model = parts[1] if len(parts) > 1 else "Unknown"
        self.serial = parts[2] if len(parts) > 2 else "Unknown"
        self.firmware = parts[3] if len(parts) > 3 else "Unknown"

    def close(self):
        pass


class InstrumentManager:
    def __init__(self, filter_mode: str = "AUTO"):
        self.filter_mode = filter_mode
        self.signal_generator = None
        self.signal_analyzer = None

    def discover(self, progress_cb=None):
        if USE_SIMULATION:
            if progress_cb:
                progress_cb("SIMULATION MODE: Found simulated instruments")
            return ["SIM::SG"], ["SIM::SA"]
        
        # Real mode (not implemented in simulation)
        return [], []

    def connect_sg(self, address: str):
        if USE_SIMULATION:
            sim_sg = SimulatedSignalGenerator()
            inst = Instrument(sim_sg, address)
            inst.manufacturer = "Keysight Technologies"
            inst.model = "N5182B (Simulated)"
            inst.serial = "SIM-SG-001"
            inst.firmware = "1.0.0"
            self.signal_generator = inst
            return inst
        return None

    def connect_sa(self, address: str):
        if USE_SIMULATION:
            sim_sa = SimulatedSignalAnalyzer()
            inst = Instrument(sim_sa, address)
            inst.manufacturer = "Keysight Technologies"
            inst.model = "N9020B (Simulated)"
            inst.serial = "SIM-SA-001"
            inst.firmware = "1.0.0"
            self.signal_analyzer = inst
            return inst
        return None

    def disconnect_all(self):
        self.signal_generator = None
        self.signal_analyzer = None


# ===========================================================================
# SCPI WRAPPERS (Simulation Mode)
# ===========================================================================

class SignalGeneratorSCPI:
    def __init__(self, inst: Instrument):
        self.inst = inst
        if USE_SIMULATION:
            self._sim = SimulatedSignalGenerator()
        else:
            self._sim = None

    def set_frequency_hz(self, freq_hz: float):
        if USE_SIMULATION:
            self._sim.set_frequency_hz(freq_hz)
        else:
            self.inst.write(f":FREQ {freq_hz:.6f}Hz")

    def set_power_dbm(self, power_dbm: float):
        if USE_SIMULATION:
            self._sim.set_power_dbm(power_dbm)
        else:
            self.inst.write(f":POW {power_dbm:.3f}dBm")

    def rf_on(self):
        if USE_SIMULATION:
            self._sim.rf_on()
        else:
            self.inst.write(":OUTP ON")

    def rf_off(self):
        if USE_SIMULATION:
            self._sim.rf_off()
        else:
            self.inst.write(":OUTP OFF")


class SignalAnalyzerSCPI:
    def __init__(self, inst: Instrument):
        self.inst = inst
        if USE_SIMULATION:
            self._sim = SimulatedSignalAnalyzer()
        else:
            self._sim = None

    def set_frequency_hz(self, freq_hz: float):
        if USE_SIMULATION:
            self._sim.set_frequency_hz(freq_hz)
        else:
            self.inst.write(f":SENS:FREQ:CENT {freq_hz:.6f}Hz")

    def set_reference_level(self, ref_dbm: float):
        if USE_SIMULATION:
            self._sim.set_reference_level(ref_dbm)
        else:
            self.inst.write(f":DISP:WIND:TRAC:Y:RLEV {ref_dbm:.2f}dBm")

    def auto_tune(self) -> dict:
        if USE_SIMULATION:
            return self._sim.auto_tune()
        return {}

    def peak_search(self) -> tuple:
        if USE_SIMULATION:
            return self._sim.peak_search()
        return (0, 0)

    def set_input_power(self, power_dbm: float):
        """Helper for simulation to track input power."""
        if USE_SIMULATION and hasattr(self._sim, 'set_input_power'):
            self._sim.set_input_power(power_dbm)

    def read_channel_power(self) -> float:
        if USE_SIMULATION:
            return self._sim.read_channel_power()
        return 0


# ===========================================================================
# EXCEL REPORTER
# ===========================================================================

class ExcelReporter:
    OUTPUT_ROOT = Path("./P1dB_Test_Results")
    
    def __init__(self):
        self.wb = None
        self.ws = None
        self.filepath = None
        self._row = 1
        self._data_start_row = 1

    def create(self, meta: dict) -> Path:
        if openpyxl is None:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")
        
        # Create folder
        proj = self._sanitize(meta.get("project_name", "Project"))
        unit = self._sanitize(meta.get("unit_name", "Unit"))
        cond = self._sanitize(meta.get("test_condition", "Condition"))
        folder = self.OUTPUT_ROOT / f"{proj}_{unit}_{cond}"
        folder.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filepath = folder / f"P1dB_Test_{timestamp}.xlsx"
        
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "P1dB Results"
        
        self._write_header(meta)
        self._write_column_headers()
        self.wb.save(self.filepath)
        
        return self.filepath

    @staticmethod
    def _sanitize(s: str) -> str:
        return "".join(c for c in s if c.isalnum() or c in ("_", "-"))[:40]

    def _write_header(self, meta: dict):
        ws = self.ws
        fields = [
            ("Project Name", meta.get("project_name", "")),
            ("Unit Name", meta.get("unit_name", "")),
            ("Test Condition", meta.get("test_condition", "")),
            ("Date", datetime.now().strftime("%Y-%m-%d")),
            ("Time", datetime.now().strftime("%H:%M:%S")),
            ("Center Frequency", f"{meta.get('center_freq_mhz', 2450)} MHz"),
            ("Reference Level", f"{meta.get('ref_level', -10)} dBm"),
            ("Measurement Type", meta.get("meas_type", "Marker Power")),
            ("SIMULATION MODE", "ACTIVE - Test Data"),
        ]
        
        row = 1
        for label, value in fields:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=str(value))
            row += 1
        
        self._row = row + 1

    def _write_column_headers(self):
        ws = self.ws
        headers = ["Timestamp", "Input Power (dBm)", "Output Power (dBm)",
                   "Gain (dB)", "Compression (dB)"]
        
        for i, header in enumerate(headers):
            cell = ws.cell(row=self._row, column=i+4, value=header)
            cell.font = Font(bold=True)
        
        self._data_start_row = self._row + 1
        self._row = self._data_start_row

    def append_row(self, timestamp: str, in_pwr: float, out_pwr: float,
                   gain: float, compression: float, highlight: bool = False):
        if self.ws is None:
            return
        
        row = self._row
        values = [timestamp, round(in_pwr, 3), round(out_pwr, 3),
                  round(gain, 3), round(compression, 3)]
        
        for i, value in enumerate(values):
            cell = self.ws.cell(row=row, column=i+4, value=value)
            if highlight:
                cell.fill = PatternFill("solid", fgColor="00FF00")
        
        self._row += 1
        self.wb.save(self.filepath)

    def save(self):
        if self.wb and self.filepath:
            self.wb.save(self.filepath)


# ===========================================================================
# TEST ENGINE QThread (FIXED - removed method name conflict)
# ===========================================================================

class TestEngine(QThread):
    sig_log = pyqtSignal(str)
    sig_status = pyqtSignal(str)
    sig_progress = pyqtSignal(int, int)
    sig_measurement = pyqtSignal(dict)
    sig_p1db_found = pyqtSignal(dict)
    sig_error = pyqtSignal(str)
    sig_finished = pyqtSignal()
    sig_tune_data = pyqtSignal(dict)
    sig_excel_saved = pyqtSignal(str)

    def __init__(self, sg: SignalGeneratorSCPI, sa: SignalAnalyzerSCPI,
                 params: dict, reporter: ExcelReporter):
        super().__init__()
        self.sg = sg
        self.sa = sa
        self.params = params
        self.reporter = reporter
        self.stop_requested = False  # Renamed from _stop_flag to avoid conflict
        self._mutex = QMutex()

    def stop(self):
        with QMutexLocker(self._mutex):
            self.stop_requested = True

    def is_stopped(self) -> bool:
        with QMutexLocker(self._mutex):
            return self.stop_requested

    def run(self):
        p = self.params
        
        try:
            # Configure instruments
            self.sig_log.emit("Configuring instruments...")
            self.sg.set_frequency_hz(p["center_freq_hz"])
            self.sa.set_frequency_hz(p["center_freq_hz"])
            self.sa.set_reference_level(p["ref_level"])
            
            # RF ON
            self.sig_log.emit("RF ON")
            self.sg.set_power_dbm(p["start_power"])
            self.sa.set_input_power(p["start_power"])  # For simulation
            self.sg.rf_on()
            time.sleep(0.5)
            
            # Auto Tune
            if p.get("auto_tune", True):
                self.sig_log.emit("Auto Tune...")
                tune_data = self.sa.auto_tune()
                self.sig_tune_data.emit(tune_data)
            
            # Reference Gain (5 samples)
            self.sig_log.emit("Measuring reference gain...")
            ref_gains = []
            for i in range(5):
                if self.is_stopped():
                    self.sg.rf_off()
                    return
                
                self.sg.set_power_dbm(p["start_power"])
                self.sa.set_input_power(p["start_power"])
                time.sleep(p["settle_time"])
                
                _, amp = self.sa.peak_search()
                out_pwr = amp + p["out_loss"]
                in_pwr = p["start_power"] - p["in_loss"]
                ref_gains.append(out_pwr - in_pwr)
            
            ref_gain = sum(ref_gains) / len(ref_gains)
            self.sig_log.emit(f"Reference Gain = {ref_gain:.2f} dB")
            self.sig_measurement.emit({"ref_gain": ref_gain})
            
            # Sweep
            steps = []
            power = p["start_power"]
            while power <= p["stop_power"] + 0.001:
                steps.append(power)
                power += p["step_size"]
            
            total = len(steps)
            self.sig_log.emit(f"Sweeping {total} points from {p['start_power']} to {p['stop_power']} dBm...")
            
            for idx, set_pwr in enumerate(steps):
                if self.is_stopped():
                    self.sig_log.emit("Test stopped by user")
                    break
                
                self.sg.set_power_dbm(set_pwr)
                self.sa.set_input_power(set_pwr)
                time.sleep(p["settle_time"])
                
                freq, amp = self.sa.peak_search()
                out_pwr = amp + p["out_loss"]
                in_pwr = set_pwr - p["in_loss"]
                gain = out_pwr - in_pwr
                compression = ref_gain - gain
                
                timestamp = datetime.now().strftime("%H:%M:%S")
                p1db_hit = compression >= 1.0
                
                # Log to Excel
                self.reporter.append_row(timestamp, in_pwr, out_pwr, 
                                         gain, compression, p1db_hit)
                
                # Emit measurement
                meas = {
                    "in_pwr": in_pwr,
                    "out_pwr": out_pwr,
                    "gain": gain,
                    "compression": compression,
                    "marker_freq": freq,
                    "marker_amp": amp,
                }
                self.sig_measurement.emit(meas)
                self.sig_progress.emit(idx + 1, total)
                self.sig_log.emit(f"Step {idx+1}/{total}: Pin={in_pwr:.2f} dBm, "
                                 f"Gain={gain:.2f} dB, Comp={compression:.3f} dB")
                
                if p1db_hit:
                    self.sig_log.emit(f"★ P1dB reached at {in_pwr:.2f} dBm input")
                    self.sig_p1db_found.emit({
                        "in_pwr": in_pwr,
                        "out_pwr": out_pwr,
                        "gain": gain,
                        "compression": compression,
                    })
                    break
            
            # Cleanup
            self.sg.rf_off()
            self.reporter.save()
            self.sig_excel_saved.emit(str(self.reporter.filepath))
            self.sig_log.emit("Test completed successfully")
            self.sig_finished.emit()
            
        except Exception as e:
            error_msg = f"Test error: {str(e)}\n{traceback.format_exc()}"
            self.sig_log.emit(f"ERROR: {error_msg}")
            self.sig_error.emit(str(e))
            try:
                self.sg.rf_off()
            except:
                pass


# ===========================================================================
# MAIN WINDOW
# ===========================================================================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("P1dB Compression Point Tester" + 
                           (" [SIMULATION MODE]" if USE_SIMULATION else ""))
        self.setGeometry(100, 100, 1200, 800)
        
        # State variables
        self.sg_scpi = None
        self.sa_scpi = None
        self.manager = None
        self.engine = None
        self.reporter = None
        
        self._build_ui()
        self._set_status("Ready - Simulation Mode" if USE_SIMULATION else "Ready")
        
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        
        # Create splitter for left and right panels
        splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(splitter)
        
        # Left panel - Controls
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setSpacing(10)
        
        # Test Information
        test_info_group = QGroupBox("Test Information")
        test_info_layout = QGridLayout(test_info_group)
        
        self.project_name = QLineEdit("Test_Project")
        self.unit_name = QLineEdit("Unit_001")
        self.test_condition = QLineEdit("25C_Room_Temp")
        
        test_info_layout.addWidget(QLabel("Project:"), 0, 0)
        test_info_layout.addWidget(self.project_name, 0, 1)
        test_info_layout.addWidget(QLabel("Unit:"), 1, 0)
        test_info_layout.addWidget(self.unit_name, 1, 1)
        test_info_layout.addWidget(QLabel("Condition:"), 2, 0)
        test_info_layout.addWidget(self.test_condition, 2, 1)
        
        left_layout.addWidget(test_info_group)
        
        # Connection
        conn_group = QGroupBox("Instrument Connection")
        conn_layout = QVBoxLayout(conn_group)
        
        self.connect_btn = QPushButton("Connect Simulated Instruments")
        self.connect_btn.clicked.connect(self._connect_instruments)
        conn_layout.addWidget(self.connect_btn)
        
        self.disconnect_btn = QPushButton("Disconnect")
        self.disconnect_btn.clicked.connect(self._disconnect_instruments)
        self.disconnect_btn.setEnabled(False)
        conn_layout.addWidget(self.disconnect_btn)
        
        # Status indicators
        status_layout = QHBoxLayout()
        self.sg_status = QLabel("⚫ SG: Not Connected")
        self.sa_status = QLabel("⚫ SA: Not Connected")
        status_layout.addWidget(self.sg_status)
        status_layout.addWidget(self.sa_status)
        conn_layout.addLayout(status_layout)
        
        left_layout.addWidget(conn_group)
        
        # Frequency Configuration
        freq_group = QGroupBox("Frequency Configuration")
        freq_layout = QGridLayout(freq_group)
        
        self.center_freq = QDoubleSpinBox()
        self.center_freq.setRange(0.1, 40000)
        self.center_freq.setValue(2450)
        self.center_freq.setSuffix(" MHz")
        
        freq_layout.addWidget(QLabel("Center Frequency:"), 0, 0)
        freq_layout.addWidget(self.center_freq, 0, 1)
        
        left_layout.addWidget(freq_group)
        
        # Power Sweep Configuration
        sweep_group = QGroupBox("Power Sweep Configuration")
        sweep_layout = QGridLayout(sweep_group)
        
        self.start_power = QDoubleSpinBox()
        self.start_power.setRange(-50, 10)
        self.start_power.setValue(-30)
        self.start_power.setSuffix(" dBm")
        
        self.stop_power = QDoubleSpinBox()
        self.stop_power.setRange(-50, 20)
        self.stop_power.setValue(5)
        self.stop_power.setSuffix(" dBm")
        
        self.step_size = QDoubleSpinBox()
        self.step_size.setRange(0.1, 5)
        self.step_size.setValue(1.0)
        self.step_size.setSuffix(" dB")
        
        self.settle_time = QDoubleSpinBox()
        self.settle_time.setRange(0.05, 2)
        self.settle_time.setValue(0.2)
        self.settle_time.setSuffix(" s")
        
        sweep_layout.addWidget(QLabel("Start Power:"), 0, 0)
        sweep_layout.addWidget(self.start_power, 0, 1)
        sweep_layout.addWidget(QLabel("Stop Power:"), 1, 0)
        sweep_layout.addWidget(self.stop_power, 1, 1)
        sweep_layout.addWidget(QLabel("Step Size:"), 2, 0)
        sweep_layout.addWidget(self.step_size, 2, 1)
        sweep_layout.addWidget(QLabel("Settle Time:"), 3, 0)
        sweep_layout.addWidget(self.settle_time, 3, 1)
        
        left_layout.addWidget(sweep_group)
        
        # Loss Configuration
        loss_group = QGroupBox("Cable Loss Configuration")
        loss_layout = QGridLayout(loss_group)
        
        self.in_loss = QDoubleSpinBox()
        self.in_loss.setRange(0, 20)
        self.in_loss.setValue(0.5)
        self.in_loss.setSuffix(" dB")
        
        self.out_loss = QDoubleSpinBox()
        self.out_loss.setRange(0, 20)
        self.out_loss.setValue(0.5)
        self.out_loss.setSuffix(" dB")
        
        loss_layout.addWidget(QLabel("Input Cable Loss:"), 0, 0)
        loss_layout.addWidget(self.in_loss, 0, 1)
        loss_layout.addWidget(QLabel("Output Cable Loss:"), 1, 0)
        loss_layout.addWidget(self.out_loss, 1, 1)
        
        left_layout.addWidget(loss_group)
        
        # Measurement Options
        meas_group = QGroupBox("Measurement Options")
        meas_layout = QVBoxLayout(meas_group)
        
        self.auto_tune = QCheckBox("Auto Tune Before Test")
        self.auto_tune.setChecked(True)
        meas_layout.addWidget(self.auto_tune)
        
        self.peak_search = QCheckBox("Peak Search at Each Step")
        self.peak_search.setChecked(True)
        meas_layout.addWidget(self.peak_search)
        
        left_layout.addWidget(meas_group)
        
        # Control Buttons
        control_group = QGroupBox("Test Control")
        control_layout = QHBoxLayout(control_group)
        
        self.start_btn = QPushButton("▶ Start Test")
        self.start_btn.clicked.connect(self._start_test)
        self.start_btn.setEnabled(False)
        
        self.stop_btn = QPushButton("■ Stop Test")
        self.stop_btn.clicked.connect(self._stop_test)
        self.stop_btn.setEnabled(False)
        
        self.rf_on_btn = QPushButton("RF ON")
        self.rf_on_btn.clicked.connect(self._rf_on)
        self.rf_on_btn.setEnabled(False)
        
        self.rf_off_btn = QPushButton("RF OFF")
        self.rf_off_btn.clicked.connect(self._rf_off)
        self.rf_off_btn.setEnabled(False)
        
        control_layout.addWidget(self.start_btn)
        control_layout.addWidget(self.stop_btn)
        control_layout.addWidget(self.rf_on_btn)
        control_layout.addWidget(self.rf_off_btn)
        
        left_layout.addWidget(control_group)
        left_layout.addStretch()
        
        splitter.addWidget(left_widget)
        
        # Right panel - Results
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        
        # Progress
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        right_layout.addWidget(self.progress_bar)
        
        # Live Measurements
        meas_display_group = QGroupBox("Live Measurements")
        meas_display_layout = QGridLayout(meas_display_group)
        
        self.input_power_label = QLabel("—")
        self.output_power_label = QLabel("—")
        self.gain_label = QLabel("—")
        self.compression_label = QLabel("—")
        self.ref_gain_label = QLabel("—")
        
        meas_display_layout.addWidget(QLabel("Input Power:"), 0, 0)
        meas_display_layout.addWidget(self.input_power_label, 0, 1)
        meas_display_layout.addWidget(QLabel("Output Power:"), 1, 0)
        meas_display_layout.addWidget(self.output_power_label, 1, 1)
        meas_display_layout.addWidget(QLabel("Gain:"), 2, 0)
        meas_display_layout.addWidget(self.gain_label, 2, 1)
        meas_display_layout.addWidget(QLabel("Compression:"), 3, 0)
        meas_display_layout.addWidget(self.compression_label, 3, 1)
        meas_display_layout.addWidget(QLabel("Reference Gain:"), 4, 0)
        meas_display_layout.addWidget(self.ref_gain_label, 4, 1)
        
        right_layout.addWidget(meas_display_group)
        
        # P1dB Result
        self.p1db_group = QGroupBox("★ P1dB Compression Point Result")
        p1db_layout = QGridLayout(self.p1db_group)
        
        self.p1db_input = QLabel("—")
        self.p1db_output = QLabel("—")
        self.p1db_gain = QLabel("—")
        self.p1db_comp = QLabel("—")
        
        p1db_layout.addWidget(QLabel("Input Power at P1dB:"), 0, 0)
        p1db_layout.addWidget(self.p1db_input, 0, 1)
        p1db_layout.addWidget(QLabel("Output Power:"), 1, 0)
        p1db_layout.addWidget(self.p1db_output, 1, 1)
        p1db_layout.addWidget(QLabel("Gain at P1dB:"), 2, 0)
        p1db_layout.addWidget(self.p1db_gain, 2, 1)
        p1db_layout.addWidget(QLabel("Compression:"), 3, 0)
        p1db_layout.addWidget(self.p1db_comp, 3, 1)
        
        self.p1db_group.setVisible(False)
        right_layout.addWidget(self.p1db_group)
        
        # Log
        log_group = QGroupBox("Activity Log")
        log_layout = QVBoxLayout(log_group)
        
        self.log_text = QTextEdit()
        self.log_text.setMaximumHeight(200)
        log_layout.addWidget(self.log_text)
        
        clear_log_btn = QPushButton("Clear Log")
        clear_log_btn.clicked.connect(self.log_text.clear)
        log_layout.addWidget(clear_log_btn)
        
        right_layout.addWidget(log_group)
        
        splitter.addWidget(right_widget)
        splitter.setSizes([500, 700])
        
        # Status bar
        self.status_bar = self.statusBar()
        
    def _set_status(self, message):
        self.status_bar.showMessage(message)
        
    def _log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")
        logger.info(message)
        
    @pyqtSlot()
    def _connect_instruments(self):
        self._log("Connecting to simulated instruments...")
        
        self.manager = InstrumentManager()
        
        # Connect SG
        sg_inst = self.manager.connect_sg("SIM::SG")
        self.sg_scpi = SignalGeneratorSCPI(sg_inst)
        
        # Connect SA
        sa_inst = self.manager.connect_sa("SIM::SA")
        self.sa_scpi = SignalAnalyzerSCPI(sa_inst)
        
        self.sg_status.setText("🟢 SG: Connected (Simulated)")
        self.sa_status.setText("🟢 SA: Connected (Simulated)")
        
        self.disconnect_btn.setEnabled(True)
        self.start_btn.setEnabled(True)
        self.rf_on_btn.setEnabled(True)
        self.connect_btn.setEnabled(False)
        
        self._log("✓ Instruments connected in simulation mode")
        self._set_status("Connected - Simulation Mode")
        
    @pyqtSlot()
    def _disconnect_instruments(self):
        if self.manager:
            self.manager.disconnect_all()
        
        self.sg_scpi = None
        self.sa_scpi = None
        
        self.sg_status.setText("⚫ SG: Not Connected")
        self.sa_status.setText("⚫ SA: Not Connected")
        
        self.disconnect_btn.setEnabled(False)
        self.start_btn.setEnabled(False)
        self.rf_on_btn.setEnabled(False)
        self.rf_off_btn.setEnabled(False)
        self.connect_btn.setEnabled(True)
        
        self._log("Disconnected from instruments")
        self._set_status("Disconnected")
        
    @pyqtSlot()
    def _rf_on(self):
        if self.sg_scpi:
            freq_hz = self.center_freq.value() * 1e6
            self.sg_scpi.set_frequency_hz(freq_hz)
            self.sg_scpi.set_power_dbm(self.start_power.value())
            self.sg_scpi.rf_on()
            self.rf_on_btn.setEnabled(False)
            self.rf_off_btn.setEnabled(True)
            self._log(f"RF ON - Frequency: {self.center_freq.value()} MHz, Power: {self.start_power.value()} dBm")
            
    @pyqtSlot()
    def _rf_off(self):
        if self.sg_scpi:
            self.sg_scpi.rf_off()
            self.rf_on_btn.setEnabled(True)
            self.rf_off_btn.setEnabled(False)
            self._log("RF OFF")
            
    @pyqtSlot()
    def _start_test(self):
        if not self.sg_scpi or not self.sa_scpi:
            QMessageBox.warning(self, "Not Connected", "Please connect instruments first")
            return
            
        # Validate inputs
        if self.start_power.value() >= self.stop_power.value():
            QMessageBox.warning(self, "Invalid Range", "Start power must be less than stop power")
            return
            
        self._log("=== Starting P1dB Test ===")
        
        # Create Excel reporter
        self.reporter = ExcelReporter()
        meta = {
            "project_name": self.project_name.text(),
            "unit_name": self.unit_name.text(),
            "test_condition": self.test_condition.text(),
            "center_freq_mhz": self.center_freq.value(),
            "ref_level": -10,
            "meas_type": "Marker Power",
        }
        
        try:
            excel_path = self.reporter.create(meta)
            self._log(f"Excel report created: {excel_path}")
        except Exception as e:
            QMessageBox.critical(self, "Excel Error", f"Cannot create report: {e}")
            return
            
        # Test parameters
        params = {
            "center_freq_hz": self.center_freq.value() * 1e6,
            "ref_level": -10.0,
            "in_loss": self.in_loss.value(),
            "out_loss": self.out_loss.value(),
            "start_power": self.start_power.value(),
            "stop_power": self.stop_power.value(),
            "step_size": self.step_size.value(),
            "settle_time": self.settle_time.value(),
            "auto_tune": self.auto_tune.isChecked(),
            "peak_search": self.peak_search.isChecked(),
        }
        
        # Start test engine
        self.engine = TestEngine(self.sg_scpi, self.sa_scpi, params, self.reporter)
        self.engine.sig_log.connect(self._log)
        self.engine.sig_status.connect(self._set_status)
        self.engine.sig_progress.connect(self._update_progress)
        self.engine.sig_measurement.connect(self._update_measurements)
        self.engine.sig_p1db_found.connect(self._on_p1db_found)
        self.engine.sig_error.connect(self._on_test_error)
        self.engine.sig_finished.connect(self._on_test_finished)
        self.engine.sig_excel_saved.connect(self._on_excel_saved)
        
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.p1db_group.setVisible(False)
        
        self.engine.start()
        
    @pyqtSlot()
    def _stop_test(self):
        if self.engine and self.engine.isRunning():
            self.engine.stop()
            self._log("Stop requested by user")
            
    @pyqtSlot(int, int)
    def _update_progress(self, current, total):
        percent = int(current / total * 100)
        self.progress_bar.setValue(percent)
        
    @pyqtSlot(dict)
    def _update_measurements(self, data):
        if "ref_gain" in data:
            self.ref_gain_label.setText(f"{data['ref_gain']:.2f} dB")
        if "in_pwr" in data:
            self.input_power_label.setText(f"{data['in_pwr']:.2f} dBm")
            self.output_power_label.setText(f"{data['out_pwr']:.2f} dBm")
            self.gain_label.setText(f"{data['gain']:.2f} dB")
            self.compression_label.setText(f"{data['compression']:.3f} dB")
            
    @pyqtSlot(dict)
    def _on_p1db_found(self, data):
        self.p1db_group.setVisible(True)
        self.p1db_input.setText(f"{data['in_pwr']:.2f} dBm")
        self.p1db_output.setText(f"{data['out_pwr']:.2f} dBm")
        self.p1db_gain.setText(f"{data['gain']:.2f} dB")
        self.p1db_comp.setText(f"{data['compression']:.3f} dB")
        
        QMessageBox.information(self, "P1dB Reached",
            f"P1dB Compression Point Detected!\n\n"
            f"Input Power: {data['in_pwr']:.2f} dBm\n"
            f"Output Power: {data['out_pwr']:.2f} dBm\n"
            f"Gain: {data['gain']:.2f} dB\n"
            f"Compression: {data['compression']:.3f} dB")
        
    @pyqtSlot(str)
    def _on_test_error(self, error):
        self._log(f"ERROR: {error}")
        QMessageBox.critical(self, "Test Error", f"Test failed: {error}")
        self._reset_controls()
        
    @pyqtSlot()
    def _on_test_finished(self):
        self._log("=== Test Completed ===")
        self._reset_controls()
        
    @pyqtSlot(str)
    def _on_excel_saved(self, path):
        self._log(f"Excel report saved: {path}")
        
    def _reset_controls(self):
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.rf_on_btn.setEnabled(True)
        self.rf_off_btn.setEnabled(False)
        
    def closeEvent(self, event):
        if self.engine and self.engine.isRunning():
            self.engine.stop()
            self.engine.wait(2000)
        if self.sg_scpi:
            try:
                self.sg_scpi.rf_off()
            except:
                pass
        event.accept()


# ===========================================================================
# MAIN
# ===========================================================================

def main():
    # Check for required packages
    if openpyxl is None:
        print("\n" + "="*60)
        print("ERROR: openpyxl is not installed!")
        print("Please run: pip install openpyxl")
        print("="*60 + "\n")
        sys.exit(1)
    
    if USE_SIMULATION:
        print("\n" + "="*60)
        print("  SIMULATION MODE ACTIVE")
        print("  Testing without real instruments")
        print("  P1dB will be detected around -5 dBm input")
        print("  Excel files will be saved to: ./P1dB_Test_Results/")
        print("="*60 + "\n")
    
    app = QApplication(sys.argv)
    app.setApplicationName("P1dB Tester")
    
    window = MainWindow()
    window.show()
    
    # Show info dialog
    if USE_SIMULATION:
        QMessageBox.information(window, "Simulation Mode",
            "SIMULATION MODE ACTIVE\n\n"
            "No real instruments are connected.\n"
            "The software will generate realistic test data.\n\n"
            "Expected Results:\n"
            "• Small signal gain: ~20 dB\n"
            "• P1dB at input: ~-5 dBm\n"
            "• Output power at P1dB: ~+15 dBm\n\n"
            "Click Connect, then Start Test to begin.\n"
            "Excel reports will be generated automatically.")
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()